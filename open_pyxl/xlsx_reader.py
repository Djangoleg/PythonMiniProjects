import openpyxl
from openpyxl import load_workbook

def read_test_xlsx():

    workbook = load_workbook(filename='xlfiles\Test.xlsx')

    worksheet = workbook.active

    START_TABLE_ROW = 4

    data = []

    for i in range(START_TABLE_ROW, worksheet.max_row + 1):

        data_row = []

        for j in range(1, worksheet.max_column + 1):
            cell_obj = worksheet.cell(row = i, column = j)

            data_row.append('' if cell_obj.value is None else cell_obj.value)

        data.append(data_row)

    return data

if __name__ == '__main__':
    data = read_test_xlsx()
    print(data)

