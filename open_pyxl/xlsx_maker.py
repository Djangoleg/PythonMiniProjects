from openpyxl import Workbook
from openpyxl.styles.borders import Border, Side, BORDER_THIN
from openpyxl.utils import get_column_letter

def generate_excel(data):

    workbook = Workbook()
    worksheet = workbook.active

    worksheet.title = "Report page 1"

    # Cells border style.
    thin_border = Border(
        left=Side(border_style=BORDER_THIN, color='00000000'),
        right=Side(border_style=BORDER_THIN, color='00000000'),
        top=Side(border_style=BORDER_THIN, color='00000000'),
        bottom=Side(border_style=BORDER_THIN, color='00000000'),
    )

    worksheet.cell(row=2, column=1, value="My favorite report")

    row = 3
    column_widths = []

    # Fill data.
    for values in data:
        row += 1
        for i, value in enumerate(values):
            worksheet.cell(row=row, column=i + 1, value=value).border = thin_border

            # Fill column_widths
            if len(column_widths) > i:
                if len(str(value)) > column_widths[i]:
                    column_widths[i] = len(str(value)) + 5
            else:
                column_widths += [len(str(value)) + 5]

    # Set cells width.
    for i, column_width in enumerate(column_widths, 1):
        worksheet.column_dimensions[get_column_letter(i)].width = column_width

    # Formula.
    column_index = len(data[0])
    column_letter = get_column_letter(column_index)
    worksheet.cell(row=row + 1, column=column_index,
                   value=f'=SUM({column_letter + str(5)}:{column_letter + str(row)})').border = thin_border

    workbook.save(filename='xlfiles\Test.xlsx')

if __name__ == '__main__':

    data = [['id', 'Name', 'phone', 'wages'],
            [1, 'John', '+19545565444', 500.05],
            [2, 'Vlasya', '+19545565445', 499.05],
            [3, 'Kostya', '+19545565446', 499.45],
            [4, 'Mary', '+19545565447', 1499.45]]

    generate_excel(data)
